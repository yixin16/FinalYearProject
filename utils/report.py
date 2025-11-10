from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter, landscape
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from io import BytesIO
from flask import send_file
import pandas as pd
from flask import current_app
import os
import datetime

# --- Enhanced Excel Report ---

def create_excel_report(df_raw, df_pivot, section_info):
    """
    Generates a multi-sheet, aesthetically formatted Excel report with conditional formatting and a summary sheet.
    """
    output = BytesIO()
    filename = f"attendance_{section_info['course_code']}_{section_info['section_name']}.xlsx".replace(' ', '_')

    # --- Data Preparation ---
    total_sessions = len(df_pivot.columns) - 2
    total_students = len(df_pivot)

    def calculate_rate(row):
        present = (row == 'present').sum()
        late = (row == 'late').sum()
        return (present + late) / total_sessions if total_sessions > 0 else 0

    student_stats = df_pivot.drop(columns=['Student ID', 'Name']).apply(calculate_rate, axis=1)
    stats_df = pd.DataFrame({
        'Student ID': df_pivot['Student ID'],
        'Name': df_pivot['Name'],
        'Attendance Rate': student_stats
    })
    
    avg_attendance_rate = stats_df['Attendance Rate'].mean()

    # --- Report Generation ---
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 1. Report Overview Sheet
        overview_data = {
            'Metric': ['Course Code', 'Course Name', 'Section', 'Report Generated On', 'Total Students', 'Total Sessions', 'Average Attendance Rate'],
            'Value': [
                section_info['course_code'],
                section_info['course_name'],
                section_info['section_name'],
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                total_students,
                total_sessions,
                f"{avg_attendance_rate:.1%}"
            ]
        }
        overview_df = pd.DataFrame(overview_data)
        overview_df.to_excel(writer, index=False, sheet_name='Report Overview')

        # 2. Summary Report Sheet
        df_pivot.to_excel(writer, index=False, sheet_name='Attendance Summary')
        
        # 3. Student Statistics Sheet
        stats_df['Attendance Rate'] = stats_df['Attendance Rate'].apply(lambda x: f"{x:.1%}")
        stats_df.to_excel(writer, index=False, sheet_name='Student Statistics')

        # 4. Raw Data Sheet
        df_raw.to_excel(writer, index=False, sheet_name='Raw Data')

        # --- Formatting ---
        # Color fills for conditional formatting
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        orange_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

        # Format Attendance Summary sheet
        summary_ws = writer.sheets['Attendance Summary']
        for row in summary_ws.iter_rows(min_row=2, min_col=3):
            for cell in row:
                if cell.value == 'present':
                    cell.fill = green_fill
                elif cell.value == 'absent':
                    cell.fill = red_fill
                elif cell.value == 'late':
                    cell.fill = orange_fill
        
        # Auto-adjust column widths and freeze panes for all sheets
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = 'A2' # Freeze top row
            for col in ws.columns:
                max_length = 0
                column = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            # Style the header for better visibility
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Special formatting for overview sheet
        overview_ws = writer.sheets['Report Overview']
        overview_ws.column_dimensions['A'].width = 25
        overview_ws.column_dimensions['B'].width = 40
        for cell in overview_ws['A']:
            cell.font = Font(bold=True)

    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# --- Enhanced PDF Report ---

# Define professional color scheme
PRIMARY_COLOR = colors.HexColor('#2C3E50') # Dark Blue
SECONDARY_COLOR = colors.HexColor('#3498DB') # Light Blue
ACCENT_COLOR_RED = colors.HexColor('#E74C3C') # Red for 'absent'
ACCENT_COLOR_ORANGE = colors.HexColor('#F39C12') # Orange for 'late'
LIGHT_GREY = colors.HexColor('#ECF0F1')

def header_footer_on_every_page(canvas, doc, section_info):
    """Draws a consistent header and footer on each page."""
    canvas.saveState()
    page_width, page_height = doc.pagesize
    
    # --- Header ---
    logo_path = os.path.join(current_app.root_path, 'static', 'logo.png')
    logo_width, logo_height = 0.5 * inch, 0.5 * inch
    
    if os.path.exists(logo_path):
        canvas.drawImage(logo_path, doc.leftMargin, page_height - doc.topMargin + 10, width=logo_width, height=logo_height)
    
    header_text = f"<font name='Helvetica-Bold' size='14'>Attendance Report</font><br/><font size='10'>{section_info['course_code']}: {section_info['course_name']} ({section_info['section_name']})</font>"
    p = Paragraph(header_text, getSampleStyleSheet()['Normal'])
    p.wrapOn(canvas, doc.width - logo_width - 10, doc.topMargin)
    p.drawOn(canvas, doc.leftMargin + logo_width + 10, page_height - doc.topMargin + 10)

    # --- Footer ---
    canvas.setFont('Helvetica', 9)
    footer_text = f"Page {canvas.getPageNumber()}"
    canvas.drawRightString(page_width - doc.rightMargin, doc.bottomMargin - 20, footer_text)
    
    canvas.restoreState()

def create_pdf_report(df_raw, df_pivot, section_info):
    """Generates a professional, multi-page PDF report with enhanced styling and content."""
    output = BytesIO()
    filename = f"attendance_{section_info['course_code']}_{section_info['section_name']}.pdf".replace(' ', '_')
    
    doc = SimpleDocTemplate(output, pagesize=landscape(letter), topMargin=0.8*inch, bottomMargin=0.8*inch)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='TitleStyle', fontSize=18, fontName='Helvetica-Bold', spaceAfter=12, textColor=PRIMARY_COLOR))
    styles.add(ParagraphStyle(name='Header', fontSize=12, fontName='Helvetica-Bold', spaceAfter=10, textColor=PRIMARY_COLOR))

    elements = []

    # --- Page 1: Summary Statistics and Student Rates ---
    elements.append(Paragraph("Overall Statistics", styles['Header']))
    stats_summary = df_raw['Status'].value_counts().reset_index()
    stats_summary.columns = ['Status', 'Count']
    stats_data = [stats_summary.columns.tolist()] + stats_summary.values.tolist()
    
    stats_table = Table(stats_data, colWidths=[2*inch, 1*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), PRIMARY_COLOR),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('TOPPADDING', (0,0), (-1,0), 12),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 0.3*inch))

    # Calculate and add Student Statistics table
    total_sessions = len(df_pivot.columns) - 2
    
    def calculate_rate(row):
        present = (row == 'present').sum()
        late = (row == 'late').sum()
        return (present + late) / total_sessions if total_sessions > 0 else 0

    student_stats = df_pivot.drop(columns=['Student ID', 'Name']).apply(calculate_rate, axis=1)
    stats_df = pd.DataFrame({
        'Student ID': df_pivot['Student ID'],
        'Name': df_pivot['Name'],
        'Attendance Rate': (student_stats * 100).round(1).astype(str) + '%'
    })

    elements.append(Paragraph("Student Statistics", styles['Header']))
    student_data = [stats_df.columns.tolist()] + stats_df.values.tolist()
    student_table = Table(student_data, colWidths=[1.5*inch, 2.5*inch, 1.5*inch], repeatRows=1)
    student_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), PRIMARY_COLOR),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT_GREY])
    ]))
    elements.append(student_table)
    elements.append(PageBreak())

    # --- Page 2 onwards: Detailed Attendance Summary ---
    elements.append(Paragraph("Detailed Attendance Summary", styles['Header']))
    
    # Prepare data with Paragraphs for conditional formatting
    header = [Paragraph(f'<b>{col}</b>', styles['Normal']) for col in df_pivot.columns]
    data_with_styles = [header]
    for index, row in df_pivot.iterrows():
        styled_row = []
        for col_name, cell_value in row.items():
            text = str(cell_value)
            style = styles['Normal']
            if text == 'absent':
                text = f"<font color='{ACCENT_COLOR_RED.hexval()}'>absent</font>"
            elif text == 'late':
                text = f"<font color='{ACCENT_COLOR_ORANGE.hexval()}'>late</font>"
            styled_row.append(Paragraph(text, style))
        data_with_styles.append(styled_row)

    num_cols = len(df_pivot.columns)
    col_widths = [1.2*inch, 1.8*inch] + [(doc.width - 3*inch) / (num_cols - 2)] * (num_cols - 2)

    main_table = Table(data_with_styles, colWidths=col_widths, repeatRows=1)
    main_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), PRIMARY_COLOR),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke), # For header text
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT_GREY])
    ]))
    elements.append(main_table)
    
    # Build the document with the header/footer function
    doc.build(elements, onFirstPage=lambda c, d: header_footer_on_every_page(c, d, section_info),
                        onLaterPages=lambda c, d: header_footer_on_every_page(c, d, section_info))
    
    output.seek(0)
    return send_file(output, mimetype='application/pdf', as_attachment=True, download_name=filename)